#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, json
from pathlib import Path
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

BLUE='1F4E78'; PALE='D9EAF7'; GRAY='666666'

def head(ws,title,purpose):
    ws.sheet_view.showGridLines=False; ws['A1']=title; ws['A1'].font=Font(size=18,bold=True); ws['A2']=purpose; ws['A2'].alignment=Alignment(wrap_text=True); ws.freeze_panes='A6'

def table(ws,headers,rows,name):
    r0=5
    for c,h in enumerate(headers,1):
        x=ws.cell(r0,c,h); x.font=Font(bold=True,color='FFFFFF'); x.fill=PatternFill('solid',fgColor=BLUE); x.alignment=Alignment(wrap_text=True)
    data=rows or [['']*len(headers)]
    for r,row in enumerate(data,r0+1):
        for c,v in enumerate(row,1):
            x=ws.cell(r,c,v); x.alignment=Alignment(wrap_text=True,vertical='top'); x.font=Font(color=GRAY)
        ws.row_dimensions[r].height=34
    ref=f'A{r0}:{ws.cell(r0+len(data),len(headers)).coordinate}'
    t=Table(displayName=name,ref=ref); t.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2',showRowStripes=True); ws.add_table(t)
    for c in range(1,len(headers)+1): ws.column_dimensions[ws.cell(1,c).column_letter].width=24

def add(wb,title,purpose,headers,rows,name):
    ws=wb.create_sheet(title); head(ws,title,purpose); table(ws,headers,rows,name); return ws

def build(version,out):
    adv=version=='0.2'; wb=Workbook(); wb.properties.creator='dosevai.com'; wb.properties.title=f'Portable AI Harness v{version}'
    ws=wb.active; ws.title='README'; head(ws,f'Portable AI Harness v{version}','Portable, inspectable AI harness. The AI assistant is the runtime; this workbook stores operating rules and state.')
    notes=[('Fast start','Upload this workbook and ask the AI to read _MCP_META first, then guide one recurring activity.'),('Quick setup','Use defaults and configure one task. Unknown is valid.'),('Guided setup','Define goal, inputs, sources, boundaries, approvals and what should persist.'),('Current vs history','ACTIVE is current approved context. RETIRED is history.'),('Reflection','REFLECTIONS may propose improvements; material changes require explicit user approval.'),('Privacy','Use only data permitted by your AI service and organizational policy.')]
    if adv: notes += [('Testing','USER_STORIES -> TEST_CASES -> TEST_RUNS evaluates proposed behavior before promotion.'),('Traceability','ACTION_LOG records material operations; CHANGELOG contract changes; VERSION_LOG material baselines.')]
    for r,(a,b) in enumerate(notes,5): ws.cell(r,1,a).font=Font(bold=True); ws.cell(r,1).fill=PatternFill('solid',fgColor=PALE); ws.cell(r,2,b).alignment=Alignment(wrap_text=True); ws.row_dimensions[r].height=34
    ws.column_dimensions['A'].width=23; ws.column_dimensions['B'].width=90
    add(wb,'START','Choose the smallest useful path.',['Path','Use when','What happens','Approval owner'],[['Quick setup','One recurring task','Defaults + one workflow','User'],['Guided setup','Deliberate configuration','Goal/context/sources/boundaries/approvals','User'],['Research','Understand','Evidence workflow','User'],['Decide','Choose','Compare + Challenge','User'],['Plan & Track','Do','Plan + checkpoints','User'],['Custom','New pattern','Define a workflow','User']],'tblStart')
    add(wb,'ACTIVE','Current approved context only.',['ID','Category','Item','Value','Authority','Status','Effective from','Last reviewed','Source','Notes'],[['CTX-001','Harness','Onboarding','Incomplete','User','Active','','','Workbook','Complete only after acceptance'],['CTX-002','Governance','Unknown policy','Do not infer','Harness','Active','','','Workbook','Unknown is valid'],['CTX-003','Governance','Mutation policy','Material changes require user approval','Harness','Active','','','Workbook','No silent learning']],'tblActive')
    skills=[['SK-RESEARCH','Research','Gather/compare evidence','Question;sources','Scope > collect > classify > compare','Evidence synthesis','No','User if consequential','Active','1.0'],['SK-DECIDE','Decide','Compare options','Options;criteria','Frame > compare > recommend','Recommendation','No','User','Active','1.0'],['SK-PLAN','Plan & Track','Turn goal into steps','Goal;constraints','Decompose > dependencies > checkpoints','Plan','Propose only','User','Active','1.0'],['SK-CHALLENGE','Challenge','Structured dissent','Candidate','Hunter > Skeptic > Referee','Verdict','No','User if material','Active','1.0'],['SK-RECORD','Synthesize & Record','Preserve state/provenance','Outputs;approvals','Synthesize > record','Traceable record','Approved rows only','User','Active','1.0'],['SK-REFLECT','Reflect & Improve','Learn from work','Run evidence','Observe > propose > challenge','Proposed learning','Proposal only','User','Active','1.0']]
    if adv: skills += [['SK-STORY','Define User Stories','Define testable behavior','Idea','Actor > value > scope > acceptance','Story','Proposal only','User','Active','0.2'],['SK-TEST-DESIGN','Design Synthetic Tests','Create evaluation cases','Story','Smoke > edge > negative > regression','Test pack','No','Independent','Active','0.2'],['SK-TEST-RUN','Execute & Record Tests','Record actual results','Test case','Execute > observe > record','Evidence','Append only','Reviewer','Active','0.2'],['SK-VERSION','Version & Trace','Preserve lineage','Approved change','Action > change > material version','Lineage','Approved only','User','Active','0.2']]
    add(wb,'SKILLS','Reusable procedures.',['Skill ID','Name','Purpose','Inputs','Procedure','Output','Allowed mutation','Approval','Status','Version'],skills,'tblSkills')
    agents=[['AG-COORD','Coordinator','Route work','Choose the smallest workflow and keep scope/evidence/gates coherent.','SK-RESEARCH;SK-DECIDE;SK-PLAN;SK-CHALLENGE;SK-RECORD','Authorized context','Coordinated result','Propose only','User','Active','1.0'],['AG-HUNTER','Hunter','Find value','Strengthen idea; find evidence/reuse/extensions.','SK-RESEARCH;SK-CHALLENGE','Candidate','Strengthening case','None','N/A','Active','1.0'],['AG-SKEPTIC','Skeptic','Challenge','Try to falsify assumptions, claims and unnecessary complexity.','SK-RESEARCH;SK-CHALLENGE','Candidate + Hunter','Challenge case','None','N/A','Active','1.0'],['AG-REFEREE','Referee','Adjudicate','Promote, repair, combine, defer or reject.','SK-CHALLENGE;SK-RECORD','Hunter + Skeptic','Verdict','None','User if material','Active','1.0'],['AG-REFLECT','Reflector','Propose learning','Convert run evidence into the smallest reusable improvement proposal.','SK-REFLECT;SK-CHALLENGE','Run evidence','Reflection proposal','Proposal only','User','Active','1.0'],['AG-RECORDER','Recorder','Preserve lineage','Record approved state and provenance without upgrading evidence by inference.','SK-RECORD','Approved outputs','Updated records','Approved only','User','Active','1.0']]
    if adv: agents += [['AG-STORY','Story Designer','Define behavior','Create actor/value/scope/non-goals/Given-When-Then acceptance/evidence.','SK-STORY','Candidate capability','Draft story','Proposal only','User','Active','0.2'],['AG-TEST','Test Designer & Evaluator','Design/evaluate','Derive synthetic tests and record actual results; never approve the change tested.','SK-TEST-DESIGN;SK-TEST-RUN','Story + runtime','Test evidence','Append evidence only','Independent approval','Active','0.2']]
    add(wb,'AGENTS','Declarative role profiles interpreted by the external AI runtime.',['Agent ID','Name','Purpose','Instruction','Permitted skills','Input scope','Output','Mutation mode','Approval','Status','Version'],agents,'tblAgents')
    workflows=[['WF-RESEARCH','Research','Research','Question;sources','Scope > research > evidence > synthesis > reflect','Coordinator;Hunter;Skeptic;Recorder','SK-RESEARCH;SK-CHALLENGE;SK-RECORD;SK-REFLECT','Synthesis','User if consequential','Active','1.0'],['WF-DECIDE','Decide','Choose','Options;criteria','Frame > Hunter > Skeptic > Referee > record > reflect','Coordinator;Hunter;Skeptic;Referee;Recorder;Reflector','SK-DECIDE;SK-CHALLENGE;SK-RECORD;SK-REFLECT','Decision+rationale','User','Active','1.0'],['WF-PLAN','Plan & Track','Plan','Goal;constraints','Scope > plan > dependencies > checkpoints > review > reflect','Coordinator;Skeptic;Recorder;Reflector','SK-PLAN;SK-CHALLENGE;SK-RECORD;SK-REFLECT','Plan','User','Active','1.0'],['WF-REFLECT','Reflect & Improve','Reflect','Completed run','Observe > propose > Skeptic > Referee > user decision > record','Reflector;Skeptic;Referee;Recorder','SK-REFLECT;SK-CHALLENGE;SK-RECORD','Learning decision','User','Active','1.0']]
    if adv: workflows += [['WF-CHANGE','Validate Material Change','Material change','Reflection/proposal','Story > tests > execute > Skeptic > Referee > human approval > apply > trace','Story Designer;Test Designer & Evaluator;Skeptic;Referee;Recorder','SK-STORY;SK-TEST-DESIGN;SK-TEST-RUN;SK-CHALLENGE;SK-VERSION','Validated change decision','User','Active','0.2']]
    add(wb,'WORKFLOWS','Reusable sequences.',['Workflow ID','Name','Trigger','Inputs','Steps','Agents','Skills','Output','Approval','Status','Version'],workflows,'tblWorkflows')
    if adv:
        add(wb,'USER_STORIES','Define proposed behavior before promotion.',['Story ID','Actor','Need','Value','Scope','Non-goals','Dependencies','Acceptance criteria','Risk','Status','Evidence required','Version'],[['US-001','Harness user','Continue from saved workbook','Avoid rebuilding context','Read current/proposed/history/version','No autonomous execution','Workbook readable','Given fresh chat, when workbook supplied, then states are distinguished','Medium','Draft','Fresh-chat evidence','0.2'],['US-002','Harness user','Challenge material idea','Reduce confirmation bias','Hunter>Skeptic>Referee','No independent-agent claim','Challenge workflow','Given candidate, when Challenge runs, then perspectives differ and one verdict is issued','Low','Draft','Recorded output','0.2']],'tblUserStories')
        tests=[[f'TC-{i:02d}','US-001' if i<=5 else 'US-002','Smoke',f'Synthetic case {i}','Clean baseline','Behavior exercised','No state/approval boundary bypassed','TEST_RUNS row','High' if i<=5 else 'Medium','Draft','Yes','0.2'] for i in range(1,11)]
        add(wb,'TEST_CASES','Synthetic tests; about 10 smoke cases by default, expand toward 20 when justified.',['Test ID','Story ID','Type','Scenario','Given','When','Then','Expected evidence','Priority','Status','Regression','Version'],tests,'tblTestCases')
        add(wb,'TEST_RUNS','Actual evidence; never mark unexecuted tests Pass.',['Run ID','Date','Test ID','Harness version','Provider/runtime','Input fingerprint','Observed result','Status','Evidence','Failure mode','Reviewer','Decision','Next action','Notes'],[],'tblTestRuns')
        add(wb,'ACTION_LOG','Material operations/decisions; action row does not prove approval.',['Action ID','Date','Actor','Action type','Object','Before','After','Reason','Approval required','Approval state','Evidence','Result','Version','Notes'],[],'tblActionLog')
        add(wb,'VERSION_LOG','Material baselines/rollback lineage.',['Version','Date','State','Parent version','Summary','Artifact hash','Validation','Approval','Rollback target','Notes'],[['0.1','','Baseline','','Core harness','','Structural','N/A','','Historical baseline'],['0.2','','Candidate','0.1','Adds stories/tests/action/version logs','','Structural; runtime pending','Pending','0.1','Recommended experiment baseline']],'tblVersionLog')
    add(wb,'SOURCES','Optional source registry; sources never silently override current state.',['Source ID','Name','Location','Purpose','Authority','Update mode','Last checked','Status','Notes','Version'],[['SRC-001','Workbook contract','This workbook','Operating rules','Authoritative workbook-local','On demand','','Active','No external source required','1.0']],'tblSources')
    add(wb,'REFLECTIONS','Observations/proposals, not active truth until approved.',['Reflection ID','Date','Workflow','Observation','Evidence','Proposed improvement','Target object','Risk','Skeptic view','Referee verdict','User decision','Applied','Version','Notes'],[],'tblReflections')
    add(wb,'EXPERIMENTS','Exploratory comparisons, distinct from validated outcomes.',['Experiment ID','Date','Question','Baseline version','Variant','Provider/runtime','Inputs','Expected','Observed','Result','Limitations','Decision','Next step','Notes'],[],'tblExperiments')
    add(wb,'CHANGELOG','Material design/contract changes only.',['Change ID','Date','Version before','Version after','Change type','Affected object','Reason','Approval','Evidence','Notes'],[],'tblChangelog')
    add(wb,'RETIRED','Superseded history; never current unless reactivated.',['ID','Category','Item','Previous value','Retired on','Replaced by','Reason','Source'],[],'tblRetired')
    meta=[['harness_type','portable_ai_harness','text'],['schema_version',version,'semver'],['release_state','candidate' if adv else 'baseline','enum'],['onboarding_status','incomplete','enum'],['runtime_model','external_llm_or_ai_assistant','text'],['current_sheet','ACTIVE','sheet'],['history_sheet','RETIRED','sheet'],['skills_sheet','SKILLS','sheet'],['agents_sheet','AGENTS','sheet'],['workflows_sheet','WORKFLOWS','sheet'],['sources_sheet','SOURCES','sheet'],['reflections_sheet','REFLECTIONS','sheet'],['changes_sheet','CHANGELOG','sheet'],['unknown_policy','do_not_infer','enum'],['mutation_policy','material_changes_require_user_approval','enum'],['reflection_policy','proposal_only_until_approved','enum'],['agent_policy','profiles_are_declarative_not_autonomous','enum'],['provider_policy','record_differences_do_not_assume_parity','enum']]
    if adv: meta += [['user_stories_sheet','USER_STORIES','sheet'],['test_cases_sheet','TEST_CASES','sheet'],['test_runs_sheet','TEST_RUNS','sheet'],['action_log_sheet','ACTION_LOG','sheet'],['version_log_sheet','VERSION_LOG','sheet'],['default_test_depth','10_smoke_expand_to_20_when_justified','text'],['test_statuses','Pass|Fail|Blocked|Error','enum'],['segregation_rule','tester_cannot_approve_change_it_tests','text'],['material_change_flow','propose>story>tests>execute>skeptic>referee>human_approval>apply>trace>version_if_material','text']]
    add(wb,'_MCP_META','Read first; pointers and invariants only.',['Key','Value','Type'],meta,'tblMCPMeta')
    out.mkdir(parents=True,exist_ok=True); p=out/f'portable-ai-harness-v{version}.xlsx'; wb.save(p); return p

def scan(p):
    bad=[]
    with ZipFile(p) as z:
        for n in [x.lower() for x in z.namelist()]:
            if any(m in f'/{n}' for m in ('vbaproject.bin','/activex/','/embeddings/','/externallinks/','/connections.xml','/customui/')): bad.append(n)
    return {'file':p.name,'size_bytes':p.stat().st_size,'sha256':hashlib.sha256(p.read_bytes()).hexdigest(),'suspicious_parts':bad}

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--root',type=Path,default=Path('artifacts/portable-ai-harness')); a=ap.parse_args()
    for v in ('0.1','0.2'):
        folder=a.root/f'v{v}'; p=build(v,folder); e=scan(p); (folder/f'portable-ai-harness-v{v}-scan.json').write_text(json.dumps(e,indent=2)+'\n'); print(v,e['sha256'],e['size_bytes'])
    return 0
if __name__=='__main__': raise SystemExit(main())
